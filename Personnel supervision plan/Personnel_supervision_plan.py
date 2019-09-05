#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
import easygui as g
import os.path
import win32com.client as win32com
address=os.path.abspath('.')
from openpyxl.styles import Border, Side, Font 
border=Border(left=Side(style='thin'),right=Side(style='thin'),
       top=Side(style='thin'),bottom=Side(style='thin'))
wb=openpyxl.load_workbook('%s\\TC_XMN_F_09.02CS E 质量监控计划.xlsx'%address)
ws=wb.get_sheet_by_name('2019_TC_XMN_F_09.02CS E 质量监控计划')
n=4
m=5
x=2
while ws.cell(row=n,column=4).value is not None:
    peple=ws.cell(row=m,column=6).value
    print(peple)
    name=peple.split(',')
    print(name)
    test_method=ws.cell(row=m,column=4).value
    test=ws.cell(row=n,column=4).value
    schedule=ws.cell(row=n,column=5).value
    supervisor=ws.cell(row=n,column=7).value
    implemented_date=ws.cell(row=n,column=8).value
    for each in name:
        wb_supervision=openpyxl.load_workbook('%s\\TC_XMN_F_16.08CS E 人员监督计划.xlsx'%address)
        ws_supervision=wb_supervision.get_sheet_by_name('TC_XMN_F_16.08CS E 人员监督计划')    
        print(n,m,x)
        while ws_supervision.cell(row=x,column=2).value is not None:
            x+=1
        ws_supervision.cell(row=x,column=1).value=x-1
        ws_supervision.cell(row=x,column=2).value=each
        ws_supervision.cell(row=x,column=3).value='%s %s'%(test,test_method)
        ws_supervision.cell(row=x,column=4).value='人员监督/人员能力监控'
        ws_supervision.cell(row=x,column=5).value=supervisor
        ws_supervision.cell(row=x,column=6).value=schedule
        ws_supervision.cell(row=x,column=7).value=implemented_date
        wb_supervision.save('%s\\TC_XMN_F_16.08CS E 人员监督计划.xlsx'%address)
    n+=3
    m+=3

g.msgbox('All the records have been generated')


