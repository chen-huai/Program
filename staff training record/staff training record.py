#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
import easygui as g
import os.path
import win32com.client as win32com
address=os.path.abspath('.')


from openpyxl.styles import Border, Side, Font 
border=Border(left=Side(style='thin'),right=Side(style='thin'),
       top=Side(style='thin'),bottom=Side(style='thin'))
wb=openpyxl.load_workbook('%s\\TC_XMN_F_16.03CS E 年度人员培训需求计划.xlsx'%address)
ws=wb.get_sheet_by_name('Staff Training Plan')
n=2
while ws.cell(row=n,column=3).value is not None:
    training=ws.cell(row=n,column=3).value
    training_name=training.split(',')
    if '' in training_name:
        training_name.remove('')
    training_item=ws.cell(row=n,column=2).value
    training_date=ws.cell(row=n,column=5).value
    training_result=ws.cell(row=n,column=7).value
    trainer=ws.cell(row=n,column=6).value
    for each in training_name:
        if os.path.exists('%s\\%s.xlsx'%(address,each)):#找文件夹里有没有这个文件
            wb_template=openpyxl.load_workbook('%s\\%s.xlsx'%(address,each))
            ws_template=wb_template.get_sheet_by_name('Staff Training Record')
        else:
            excel0 = win32com.gencache.EnsureDispatch('Excel.Application')
            excel0.Visible = False
            excel0.Application.DisplayAlerts = False
            workbook0 = excel0.Workbooks.Open(os.path.join(os.getcwd(),'%s\\template.xlsx'%address))
            print(1)
            print(each)
            workbook0.SaveAs('%s\\%s.xlsx'%(address,each))
            print(2)
            excel0.Application.Quit()
            wb_template=openpyxl.load_workbook('%s\\%s.xlsx'%(address,each))
            ws_template=wb_template.get_sheet_by_name('Staff Training Record')
            
        m=5
        while ws_template.cell(row=m,column=2).value is not None:
            m+=1
        ws_template.cell(row=1,column=1).value='Name姓名：%s'%each
        ws_template.cell(row=m,column=2).value=training_date #培训日期
        ws_template.cell(row=m,column=3).value=training_item #培训项目
        ws_template.cell(row=m,column=4).value=training_result #培训结果
        ws_template.cell(row=m,column=5).value=trainer #培训机构
        ws_template['A1'].border=border
        ws_template['A2'].border=border
        ws_template['B1'].border=border
        ws_template['B2'].border=border
        ws_template['D1'].border=border
        ws_template['D2'].border=border
        ws_template['E1'].border=border
        ws_template['E2'].border=border
        wb_template.save('.\\temp.xlsx')
        

        #保存为有图片的
        excel = win32com.gencache.EnsureDispatch('Excel.Application')
        excel.Visible = False
        excel.Application.DisplayAlerts = False
        workbook = excel.Workbooks.Open(os.path.join(os.getcwd(),'%s\\temp.xlsx'%address))
        excel.Sheets("Staff Training Record").Select()
        excel.Range("A1:E41").Select()
        excel.Selection.Copy()


        excel2 = win32com.gencache.EnsureDispatch('Excel.Application')
        excel2.Visible = False
        excel2.Application.DisplayAlerts = False
        print(each)
        workbook2 = excel2.Workbooks.Open(os.path.join(os.getcwd(),'%s\\%s.xlsx'%(address,each)))

        excel2.Sheets("Staff Training Record").Select()
        excel2.Range("A1:E41").Select()
        excel2.ActiveSheet.Paste()


        workbook.Save()
        workbook2.Save()
        excel.Application.Quit()
        excel2.Application.Quit()

        os.remove('%s\\temp.xlsx'%address)
    n+=1

    
#重新启动“保存提醒”
excel3 = win32com.gencache.EnsureDispatch('Excel.Application')
excel3.Application.DisplayAlerts = True
excel3.Application.Quit()

wb.save('%s\\TC_XMN_F_16.03CS E  年度人员培训需求计划.xlsx'%address)
g.msgbox('All the records have been generated','staff training record')
